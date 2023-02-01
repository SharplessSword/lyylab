import math

import openpyxl
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from numpy import log
from scipy.optimize import curve_fit
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color


# def f_model(x, A, n, Ea):
#     return log(A)-n*log(1/x)- Ea/(1.9872*x)

def f_model(x, A, n, E):
    # print(x, A, n, E)
    return log(A)+n*log(x)-E/(x*1.9872)


# popt, pcov = curve_fit(f=f_model, xdata=xdata, ydata=ydata, p0=(3, -2, 1),)
def fitnlm(xdata, ydata, f_model):
    popt, pcov = curve_fit(f=f_model, xdata=xdata, ydata=ydata, maxfev = 8000)
    return popt

def get_fit_curve_ydata(A, n ,Ea, T):
    # a = math.log(A)-n*math.log(1/T)-Ea/(1.9872*T) # lnk = a
    # print('a', a)
    # k = math.e**a
    k = A*(T**n)*math.exp(-(Ea/8.3145))
    return k
def search_out_file():
    file_list = []
    for file in os.listdir():
        if file.endswith('.out'):
            file_list.append(file)
    return file_list


def extract_useful_part(file_name):
    data = []
    with open('./{}'.format(file_name)) as f:
        start = False
        end = False
        for line in f.readlines():
            if line.startswith('Temperature-Species Rate Tables:'):
                start = True
            if start and line.startswith('Reactant = P1 '):
                end = True
                break
            if start and not end:
                data.append(line)
    return data


def split_data(data):
    splited_data = []
    for i in data[2:]: # remove headers
        if i.startswith('Reactant'):
            splited_data.append([])
            # add Reactant
            splited_data[-1].append(i.strip())
        else:
            splited_data[-1].append(i)
    return splited_data


def get_data(splited_data):
    result_list = []
    for block in splited_data:
        result_list.append([])
        for line in block:
            line = line.strip()
            if line.startswith('Reactant'):
                result_list[-1].append(line)
            elif line != '':
                result_list[-1].append(line.split())

    return result_list


def calculate(result):
    calculated_result =[]
    for block in result:
        calculated_result.append([])
        for i, line in enumerate(block):
            if i == 0 :
                calculated_result[-1].append(line)
            elif i == 1:
                line_list = []
                line_list.append(line[0])
                for column_name in line[1:-2]:
                    line_list.append(column_name)
                    line_list.append('分支比')
                line_list += line[-2:]
                calculated_result[-1].append(line_list)
            else:
                line_list = []
                line_list.append(line[0])
                for column_value in line[1:-2]:
                    column_value = float(column_value)
                    line_list.append(column_value)
                    line_list.append(column_value/float(line[-2]))
                line_list += line[-2:]
                calculated_result[-1].append(line_list)
    return calculated_result


def fill_xls(result):

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'mysheet'

    for block_count, block in enumerate(result):
        for i, line in enumerate(block):
            if i == 0:
                worksheet.cell(block_count*len(block)+i+1, 1, line)
            else:
                for column_count, column_item in enumerate(line):
                    worksheet.cell(block_count*len(block)+i+1, column_count+1, column_item)
    block_count = len(result)
    line_count = len(result[0])
    column_count = len(result[0][1])
    return workbook, block_count, line_count, column_count

    # workbook.save(filename='1.xls')


def beautify(workbook, block_count, line_cont, column_count):
    # 每块标题美化
    ws = workbook.active
    title_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    title_fill = PatternFill(fill_type='solid', fgColor='4F81BD')
    title_alignment = Alignment(horizontal='center', vertical='center')
    for i in range(1, block_count * line_cont, line_cont):
        ws['A{}'.format(i)].font = title_font
        ws['A{}'.format(i)].fill = title_fill
        ws['A{}'.format(i)].alignment = title_alignment
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=column_count)
        ws.row_dimensions[i].height = 40
    # 列名美化
    column_name_font = Font(name='Arial', size=11)
    column_name_fill = PatternFill(fill_type='solid', fgColor='B8CCE4')
    column_name_aligment = Alignment(horizontal='center')
    for i in range(2, block_count * line_cont, line_cont):
        for column in range(1,column_count+1):
            selected_cell = ws.cell(i, column)
            selected_cell.font = column_name_font
            selected_cell.fill = column_name_fill
            selected_cell.alignment = column_name_aligment
    # 数值美化
    value_font = Font(name='Arial', size=11)
    value_fill1 = PatternFill(fill_type='solid', fgColor='DCE6F1')
    value_fill2 = PatternFill(fill_type='solid', fgColor='B8CCE4')
    vaule_boarder = Border(#left=Side(border_style='thin', color='000000'),
                           right=Side(border_style='thin', color='000000'),
                           #top=Side(border_style='thin', color='000000'),
                           bottom=Side(border_style='thin', color='000000'))
    for i in range(3, block_count * line_cont, line_cont):
        for row in range(0, line_cont - 2):
            for column in range(1, column_count+1):
                selected_cell = ws.cell(i+row, column)
                # print(selected_cell.row, selected_cell.column)
                selected_cell.border = vaule_boarder
                selected_cell.font = value_font
                if row % 2 == 0:
                    selected_cell.fill = value_fill1
                else:
                    selected_cell.fill = value_fill2
    # 数值部分颜色分层
    for i in range(3, block_count * line_cont, line_cont):
        for row in range(line_cont-2):
            for column in range(1, column_count+1):
                selected_cell = ws.cell(i+row, column)
                # print(selected_cell.row, selected_cell.column)
                selected_cell.border = vaule_boarder
                if row % 2 == 0:
                    selected_cell.fill = value_fill1
                else:
                    selected_cell.fill = value_fill2
    # 高亮显示数值大于0.1的

    highlight_value_font = Font(name='Arial', size=11, bold=True)
    highlight_value_fill = PatternFill(fill_type='solid', fgColor='00B0F0')
    for i in range(3, block_count * line_cont, line_cont):
        for row in range(line_cont-2):
            for column in range(1, column_count+1):
                if 3<=column<=column_count-2 and column%2 == 1:
                    selected_cell = ws.cell(i+row, column)
                    if selected_cell.value >0.1:
                        selected_cell.font = highlight_value_font
                        selected_cell.fill = highlight_value_fill
    return workbook


def save_xls(workbook, file_name):
    file_name = file_name.rsplit('.', 1)[0]
    print(file_name)
    workbook.save(filename='{}.xlsx'.format(file_name))


def select_model_data(result):
    model_data_tag = []
    for i, block  in enumerate(result):
        model_data_tag.append((i, 0))
        for j, line in enumerate(block):
            if j >= 2:
                for k, column in enumerate(line):
                    if 0 < k < len(line) - 2:
                        if k%2==0 and line[k]>=0.1:
                            model_data_tag.append((i, k-1))

    model_data_tag = sorted(set((model_data_tag)))
    # print(model_data_tag)
    model_data_dict = {}
    for tag in model_data_tag:
        block_index, column_index = tag
        column = [line[column_index] for line in result[block_index][1:]]
        if block_index not in model_data_dict.keys():
            model_data_dict[block_index] = {}
            model_data_dict[block_index]['data'] = [column]
        else:
            model_data_dict[block_index]['data'].append(column)
    for i, block in enumerate(result):
        model_data_dict[i]['header'] = block[0]
    return model_data_dict


def calculate_model_fitnlm(model_data_dict):
    tempurature = model_data_dict[0]['data'][0][1:]
    tempurature = [eval(i) for i in tempurature]
    for key in model_data_dict:
        model_data_dict[key]['a_n_e'] = []
        for i, column in enumerate(model_data_dict[key]['data'][1:]):
            column = column[1:]  # 去除副标题 ,如T(K)
            # print(column)
            print(column)
            ln_column = [log(i) for i in column]
            print(tempurature)
            # print(column)
            fitnlm_result = fitnlm(xdata=tempurature, ydata=ln_column, f_model=f_model)
            model_data_dict[key]['a_n_e'].append(fitnlm_result)
        # print(model_data_dict[key]['data'])
    return model_data_dict

def fill_model_xls(model_data_dict):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'mysheet'
    block_linecount = len(model_data_dict[0]['data'][0])+1
    print(block_linecount)
    for key in model_data_dict.keys():
        header = model_data_dict[key]['header']
        worksheet.cell(key*block_linecount+1, 1, header)
        for i, one_column in enumerate(model_data_dict[key]['data']):
            if i == 0:
                for j, value in enumerate(one_column):
                    worksheet.cell(block_linecount*key+j+2, i+1, value)
            else:
                for j, value in enumerate(one_column):
                    worksheet.cell(block_linecount*key+j+2, 2*i, value)
        for i, a_n_e in enumerate(model_data_dict[key]['a_n_e']):
            A, n, E = a_n_e
            worksheet.cell(block_linecount * key + 3, 2*i+3, A)
            worksheet.cell(block_linecount * key + 4, 2*i+3, n)
            worksheet.cell(block_linecount * key + 5, 2*i+3, E)
    file_name = 'test'
    workbook.save(filename='{}.xlsx'.format(file_name))
    return workbook



# file_list = search_out_file()
# for file_name in file_list:
#     print(file_name)
#     data = extract_useful_part(file_name)
#     splited_data = split_data(data)
#     result = get_data(splited_data)
#     result = calculate(result)
#     workbook, block_count, line_count, column_count = fill_xls(result)
#     workbook = beautify(workbook, block_count, line_count, column_count)
#     save_xls(workbook, file_name)
    # model_data_dict = select_model_data(result)
    # model_data_dict = calculate_model_fitnlm(model_data_dict)
    # fill_model_xls(model_data_dict)
